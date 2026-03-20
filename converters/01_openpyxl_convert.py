"""
変換ツール01: openpyxl
セルを直接走査し、結合情報・値を読み取ってMarkdownテーブルを生成する。
方眼紙Excelへの対応として結合セルの展開処理を実装。
"""

import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_FILE = Path("test_data/houganshi_sample.xlsx")
OUTPUT_DIR = Path("output/01_openpyxl")


def get_merged_cell_map(ws):
    """結合セルの情報を {(row, col): (value, is_top_left)} のマップで返す"""
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = (merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(row, col)] = top_left
    return merged_map


def cell_value(ws, row, col, merged_map):
    """結合セルを考慮してセルの値を返す"""
    key = (row, col)
    if key in merged_map:
        tl = merged_map[key]
        return ws.cell(row=tl[0], column=tl[1]).value
    return ws.cell(row=row, column=col).value


def is_row_empty(ws, row, max_col, merged_map):
    return all(
        not str(cell_value(ws, row, c, merged_map) or "").strip()
        for c in range(1, max_col + 1)
    )


def sheet_to_markdown(ws) -> str:
    """1シートをMarkdownに変換する"""
    merged_map = get_merged_cell_map(ws)

    # 実際に使われている範囲を取得
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row is None or max_col is None:
        return "*（空シート）*\n"

    # 空行をスキップしながらテーブルブロックを抽出
    lines = []
    lines.append(f"## シート: {ws.title}\n")

    # --- 全セルをCSV的に読み取り ---
    table_rows = []
    for row in range(1, max_row + 1):
        if is_row_empty(ws, row, max_col, merged_map):
            continue
        row_vals = []
        for col in range(1, max_col + 1):
            v = cell_value(ws, row, col, merged_map)
            row_vals.append(str(v).strip() if v is not None else "")
        # 末尾の空セルを除去
        while row_vals and row_vals[-1] == "":
            row_vals.pop()
        if row_vals:
            table_rows.append(row_vals)

    if not table_rows:
        return lines[0] + "*（データなし）*\n"

    # 列数を揃える
    max_cols = max(len(r) for r in table_rows)
    for r in table_rows:
        while len(r) < max_cols:
            r.append("")

    # Markdownテーブルとして出力
    # 1行目をヘッダー扱い
    header = table_rows[0]
    sep    = ["---"] * len(header)
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(sep)    + " |")
    for row in table_rows[1:]:
        lines.append("| " + " | ".join(row) + " |")

    lines.append("")
    return "\n".join(lines)


def convert(input_file: Path, output_dir: Path):
    start = time.perf_counter()
    wb = load_workbook(input_file, data_only=True)
    result_parts = [f"# openpyxl 変換結果\n\n**入力ファイル:** `{input_file}`\n"]

    for ws in wb.worksheets:
        result_parts.append(sheet_to_markdown(ws))

    elapsed = time.perf_counter() - start
    result_parts.append(f"\n---\n*変換時間: {elapsed:.3f}秒 | ツール: openpyxl {__import__('openpyxl').__version__}*\n")

    md_text = "\n".join(result_parts)
    out_path = output_dir / "result.md"
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path.write_text(md_text, encoding="utf-8")

    print(f"[01_openpyxl] 完了 ({elapsed:.3f}秒) -> {out_path}")
    return elapsed, md_text


if __name__ == "__main__":
    convert(INPUT_FILE, OUTPUT_DIR)
