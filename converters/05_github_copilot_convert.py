"""
変換ツール05: GitHub Models API (azure-ai-inference SDK)

2つのアプローチでLLMを使ってExcel→Markdownを試みる。

【アプローチA: テキストモード】
  openpyxl でセル値を抽出 → 構造化テキストとして LLM に渡す
  → LLM が意味のある Markdown に整形

【アプローチB: ビジョンモード】
  openpyxl + PIL で各シートをグリッド画像にレンダリング
  → Vision LLM (gpt-4o) に画像で渡す
  → LLM がビジュアルから Markdown を生成

必要な環境変数:
  GITHUB_TOKEN : GitHub Personal Access Token (models:read 権限)

モデル:
  テキスト: openai/gpt-4o-mini (低レートリミット・無料枠)
  ビジョン: openai/gpt-4o      (高レートリミット・無料枠)
"""

import base64
import io
import os
import time
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

INPUT_FILE = Path("test_data/houganshi_sample.xlsx")
OUTPUT_DIR = Path("output/05_github_copilot")

ENDPOINT   = "https://models.inference.ai.azure.com"
TEXT_MODEL = "openai/gpt-4o-mini"
VISION_MODEL = "openai/gpt-4o"

# シートをPNG化するときのスケール設定
CELL_PX_W = 14   # 1セル幅(px)
CELL_PX_H = 18   # 1セル高(px)
MAX_COLS  = 50   # 描画する最大列数
MAX_ROWS  = 60   # 描画する最大行数


# ---------------------------------------------------------------------------
# 共通: GitHub Models クライアント生成
# ---------------------------------------------------------------------------

def get_client():
    from azure.ai.inference import ChatCompletionsClient
    from azure.core.credentials import AzureKeyCredential

    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        raise EnvironmentError(
            "環境変数 GITHUB_TOKEN が設定されていません。\n"
            "https://github.com/settings/tokens で PAT を作成し、\n"
            "  set GITHUB_TOKEN=ghp_xxxx  (Windows)\n"
            "  export GITHUB_TOKEN=ghp_xxxx  (Mac/Linux)\n"
            "を実行してから再試行してください。"
        )
    return ChatCompletionsClient(
        endpoint=ENDPOINT,
        credential=AzureKeyCredential(token),
    )


# ---------------------------------------------------------------------------
# アプローチA: テキストモード
# ---------------------------------------------------------------------------

def extract_text_grid(ws) -> str:
    """シートの値をタブ区切りグリッドテキストとして抽出（結合セル展開済み）"""
    # 結合セルマップ
    merged_map = {}
    for rng in ws.merged_cells.ranges:
        val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_map[(r, c)] = val

    lines = []
    for row in range(1, min(ws.max_row or 0, MAX_ROWS) + 1):
        row_vals = []
        for col in range(1, min(ws.max_column or 0, MAX_COLS) + 1):
            key = (row, col)
            v = merged_map.get(key, ws.cell(row, col).value)
            row_vals.append(str(v).strip() if v is not None else "")
        # 末尾の空セルを除去
        while row_vals and row_vals[-1] == "":
            row_vals.pop()
        if any(row_vals):
            lines.append("\t".join(row_vals))
    return "\n".join(lines)


def convert_text_mode(wb, client) -> str:
    """全シートをテキストグリッドにしてLLMへ送り、Markdown化してもらう"""
    from azure.ai.inference.models import SystemMessage, UserMessage

    all_md = []
    for ws in wb.worksheets:
        grid_text = extract_text_grid(ws)
        if not grid_text.strip():
            continue

        prompt = f"""以下は「{ws.title}」というExcelシートのセルデータです（タブ区切り）。
方眼紙スタイルのExcelで、セル結合を多用したレイアウトになっています。
このデータを、内容が伝わるMarkdown形式に変換してください。

変換のポイント：
- 繰り返している同じ値はタイトル行や見出しとして認識する
- セクション見出しには ## や ### を使う
- データテーブルはMarkdownテーブルで表現する（列数は適切に絞る）
- 空白行や無意味な繰り返しは除去する
- 画像・図は「[図: ...]」と記載する

シートデータ：
```
{grid_text}
```

Markdownのみを出力してください。説明文は不要です。"""

        response = client.complete(
            model=TEXT_MODEL,
            messages=[
                SystemMessage("あなたはExcelドキュメントをMarkdownに変換する専門家です。"),
                UserMessage(prompt),
            ],
            max_tokens=4096,
        )
        md = response.choices[0].message.content
        all_md.append(f"## シート: {ws.title}\n\n{md}")

    return "\n\n---\n\n".join(all_md)


# ---------------------------------------------------------------------------
# アプローチB: ビジョンモード
# ---------------------------------------------------------------------------

def render_sheet_to_png(ws) -> io.BytesIO:
    """openpyxl + PIL でシートを簡易グリッド画像にレンダリング"""
    # フォント
    font_candidates = [
        "C:/Windows/Fonts/meiryo.ttc",
        "C:/Windows/Fonts/YuGothR.ttc",
        "C:/Windows/Fonts/msgothic.ttc",
    ]
    font = None
    for fc in font_candidates:
        try:
            font = ImageFont.truetype(fc, 10)
            break
        except (IOError, OSError):
            continue
    if font is None:
        font = ImageFont.load_default()

    # 結合セルマップ
    merged_map = {}
    merged_ranges = {}  # (r,c) -> (r1,c1,r2,c2) for top-left only
    for rng in ws.merged_cells.ranges:
        val = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_map[(r, c)] = val
        merged_ranges[(rng.min_row, rng.min_col)] = (
            rng.min_row, rng.min_col, rng.max_row, rng.max_col
        )

    max_r = min(ws.max_row or 1, MAX_ROWS)
    max_c = min(ws.max_column or 1, MAX_COLS)

    img_w = max_c * CELL_PX_W + 1
    img_h = max_r * CELL_PX_H + 1
    img = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(img)

    # グリッド線
    for c in range(max_c + 1):
        x = c * CELL_PX_W
        draw.line([(x, 0), (x, img_h)], fill=(200, 200, 200), width=1)
    for r in range(max_r + 1):
        y = r * CELL_PX_H
        draw.line([(0, y), (img_w, y)], fill=(200, 200, 200), width=1)

    # セルテキスト（重複描画を避けるため top-left のみ描画）
    drawn = set()
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            key = (r, c)
            val = merged_map.get(key, ws.cell(r, c).value)
            if val is None:
                continue
            # 結合セルの左上のみテキスト描画
            if key in merged_map and key not in merged_ranges:
                continue
            if key in drawn:
                continue
            drawn.add(key)

            x = (c - 1) * CELL_PX_W + 2
            y = (r - 1) * CELL_PX_H + 2
            text = str(val)[:12]  # 長すぎる場合は切り詰め
            draw.text((x, y), text, fill=(20, 20, 20), font=font)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def convert_vision_mode(wb, client) -> str:
    """各シートをPNG化してVision LLMへ送り、Markdown化してもらう"""
    from azure.ai.inference.models import (
        ImageContentItem, ImageUrl, SystemMessage,
        TextContentItem, UserMessage,
    )

    all_md = []
    for ws in wb.worksheets:
        png_buf = render_sheet_to_png(ws)
        png_bytes = png_buf.read()
        b64 = base64.b64encode(png_bytes).decode("utf-8")

        prompt = (
            f"これはExcelの「{ws.title}」シートを画像化したものです。"
            "方眼紙スタイルのExcelで、セル結合を多用しています。\n"
            "この画像から読み取れる内容を、Markdown形式で整理してください。\n"
            "- 見出しは ## や ### を使う\n"
            "- データテーブルはMarkdownテーブルで表現する\n"
            "- 図・画像エリアは [図: 説明] と記載\n"
            "Markdownのみを出力してください。"
        )

        response = client.complete(
            model=VISION_MODEL,
            messages=[
                SystemMessage("あなたはExcel画像からMarkdownを生成する専門家です。"),
                UserMessage([
                    TextContentItem(text=prompt),
                    ImageContentItem(
                        image_url=ImageUrl(
                            url=f"data:image/png;base64,{b64}",
                            detail="high",
                        )
                    ),
                ]),
            ],
            max_tokens=4096,
        )
        md = response.choices[0].message.content
        all_md.append(f"## シート: {ws.title}\n\n{md}")

    return "\n\n---\n\n".join(all_md)


# ---------------------------------------------------------------------------
# メイン変換関数
# ---------------------------------------------------------------------------

def convert(input_file: Path, output_dir: Path):
    start = time.perf_counter()

    client = get_client()
    wb = openpyxl.load_workbook(input_file, data_only=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    results = {}

    # --- アプローチA: テキストモード ---
    print("  [A] テキストモード変換中...")
    t0 = time.perf_counter()
    try:
        md_text = convert_text_mode(wb, client)
        elapsed_a = time.perf_counter() - t0
        results["text"] = (elapsed_a, md_text)
        out_a = output_dir / "result_text.md"
        header_a = (
            f"# GitHub Models 変換結果（テキストモード）\n\n"
            f"**モデル:** `{TEXT_MODEL}`  \n"
            f"**入力ファイル:** `{input_file}`  \n"
            f"**方法:** openpyxl でセル抽出 → LLM で Markdown 整形\n\n"
        )
        footer_a = f"\n---\n*変換時間: {elapsed_a:.3f}秒*\n"
        out_a.write_text(header_a + md_text + footer_a, encoding="utf-8")
        print(f"  [A] 完了 ({elapsed_a:.3f}秒) -> {out_a}")
    except Exception as e:
        print(f"  [A] エラー: {e}")
        elapsed_a = time.perf_counter() - t0
        results["text"] = (elapsed_a, f"エラー: {e}")

    # --- アプローチB: ビジョンモード ---
    print("  [B] ビジョンモード変換中...")
    t0 = time.perf_counter()
    try:
        md_vision = convert_vision_mode(wb, client)
        elapsed_b = time.perf_counter() - t0
        results["vision"] = (elapsed_b, md_vision)
        out_b = output_dir / "result_vision.md"
        header_b = (
            f"# GitHub Models 変換結果（ビジョンモード）\n\n"
            f"**モデル:** `{VISION_MODEL}`  \n"
            f"**入力ファイル:** `{input_file}`  \n"
            f"**方法:** openpyxl + PIL でシートをPNG化 → Vision LLM で認識\n\n"
        )
        footer_b = f"\n---\n*変換時間: {elapsed_b:.3f}秒*\n"
        out_b.write_text(header_b + md_vision + footer_b, encoding="utf-8")
        print(f"  [B] 完了 ({elapsed_b:.3f}秒) -> {out_b}")
    except Exception as e:
        print(f"  [B] エラー: {e}")
        elapsed_b = time.perf_counter() - t0
        results["vision"] = (elapsed_b, f"エラー: {e}")

    elapsed_total = time.perf_counter() - start

    # convert() の返り値は run_all.py の形式に合わせて (elapsed, md_text)
    best_md = results.get("text", (0, ""))[1]
    print(f"[05_github_copilot] 合計 {elapsed_total:.3f}秒")
    return elapsed_total, best_md


if __name__ == "__main__":
    convert(INPUT_FILE, OUTPUT_DIR)
