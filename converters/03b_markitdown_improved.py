"""
変換ツール03b: markitdown 改善版

問題: markitdown は内部で pandas.read_excel() を使うため、
      結合セルは先頭セル以外 NaN になる。

改善策: openpyxl で結合セルを事前展開してから markitdown に渡す。
       → PR #1165 (fill_merged_cells) が未マージのため、この前処理で代替する。

なお markitdown v0.1.5 時点では：
  - ExcelConverter のオプションはほぼなし
  - LLM連携オプション（llm_client）はある（今回は未使用）
  - 特定シートのみ変換・header行指定は不可
"""

import io
import time
from pathlib import Path

import openpyxl
from markitdown import MarkItDown

INPUT_FILE = Path("test_data/houganshi_sample.xlsx")
OUTPUT_DIR = Path("output/03b_markitdown_improved")


def expand_merged_cells_to_stream(input_path: Path) -> io.BytesIO:
    """
    結合セルの値をすべての結合範囲セルにコピーし、結合を解除する。
    ファイルには書き出さず BytesIO で返す（一時ファイル不要）。
    """
    wb = openpyxl.load_workbook(input_path, data_only=True)
    for ws in wb.worksheets:
        # merged_cells は走査中に変化するのでコピーしてから処理
        for merge_range in list(ws.merged_cells.ranges):
            # 左上セルの値を取得
            top_left_value = ws.cell(
                row=merge_range.min_row,
                column=merge_range.min_col
            ).value
            # 結合を解除してから全セルに値をセット
            ws.unmerge_cells(str(merge_range))
            for row in ws.iter_rows(
                min_row=merge_range.min_row,
                max_row=merge_range.max_row,
                min_col=merge_range.min_col,
                max_col=merge_range.max_col,
            ):
                for cell in row:
                    cell.value = top_left_value

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def convert(input_file: Path, output_dir: Path):
    start = time.perf_counter()

    # Step1: 結合セルを展開
    expanded_buf = expand_merged_cells_to_stream(input_file)

    # Step2: 展開済みデータを markitdown に渡す
    # markitdown は file-like object にも対応している
    md_converter = MarkItDown()
    result = md_converter.convert(expanded_buf, file_extension=".xlsx")
    md_text_raw = result.text_content

    elapsed = time.perf_counter() - start

    import markitdown
    footer = (
        f"\n---\n"
        f"*変換時間: {elapsed:.3f}秒 | "
        f"ツール: markitdown {markitdown.__version__} + openpyxl結合セル展開前処理*\n"
    )
    full_text = (
        f"# markitdown 改善版 変換結果\n\n"
        f"**入力ファイル:** `{input_file}`  \n"
        f"**改善内容:** openpyxl で結合セルを事前展開してから変換\n\n"
        + md_text_raw
        + footer
    )

    out_path = output_dir / "result.md"
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path.write_text(full_text, encoding="utf-8")

    print(f"[03b_markitdown_improved] 完了 ({elapsed:.3f}秒) -> {out_path}")
    return elapsed, full_text


if __name__ == "__main__":
    convert(INPUT_FILE, OUTPUT_DIR)
