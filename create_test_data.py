"""
方眼紙スタイルのテスト用Excelファイルを生成するスクリプト

方眼紙Excel の特徴を再現:
  - 全セルを正方形に近いサイズに統一（列幅・行高さ）
  - セル結合でレイアウトを構成
  - 罫線でブロックを区切る
  - テキスト・数値・日付など混在
  - 画像（図）の埋め込み
  - グラフの埋め込み
  - 複数シート
"""

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------
OUTPUT_DIR = Path("test_data")
OUTPUT_FILE = OUTPUT_DIR / "houganshi_sample.xlsx"

# 方眼紙設定: 列幅(文字単位)・行高さ(pt)
CELL_COL_WIDTH = 3.0   # 狭い列幅
CELL_ROW_HEIGHT = 18   # 行高さ(pt)

# 罫線スタイル
THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MEDIUM_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

# 色
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
LABEL_FILL = PatternFill("solid", fgColor="EDEDED")
TITLE_FILL = PatternFill("solid", fgColor="2E4057")


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------

def set_grid(ws, max_col: int = 60, max_row: int = 80):
    """シート全体を方眼紙グリッドに設定する"""
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = CELL_COL_WIDTH
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = CELL_ROW_HEIGHT


def write_cell(ws, row, col, value, font=None, fill=None, border=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    return cell


def merge_and_write(ws, row1, col1, row2, col2, value,
                    font=None, fill=None, border=None, alignment=None):
    """セルを結合して値を書き込む"""
    ws.merge_cells(start_row=row1, start_column=col1,
                   end_row=row2, end_column=col2)
    cell = ws.cell(row=row1, column=col1, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    # 結合範囲の全セルに罫線を適用
    if border:
        for r in range(row1, row2 + 1):
            for c in range(col1, col2 + 1):
                ws.cell(row=r, column=c).border = border
    return cell


def apply_border_range(ws, row1, col1, row2, col2, border):
    """範囲全体に同じ罫線を適用する"""
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = border


# ---------------------------------------------------------------------------
# PIL で説明用画像を生成する（外部ファイル不要）
# ---------------------------------------------------------------------------

def create_dummy_image(width=300, height=150, text="工事概要図") -> io.BytesIO:
    """方眼紙Excelに貼り付けるダミー画像を生成する"""
    img = Image.new("RGB", (width, height), color=(230, 240, 255))
    draw = ImageDraw.Draw(img)

    # 格子線
    for x in range(0, width, 30):
        draw.line([(x, 0), (x, height)], fill=(180, 200, 230), width=1)
    for y in range(0, height, 30):
        draw.line([(0, y), (width, y)], fill=(180, 200, 230), width=1)

    # 簡易な図形（建物風）
    draw.rectangle([60, 50, 140, 120], outline=(50, 80, 150), width=2, fill=(200, 215, 240))
    draw.rectangle([160, 70, 240, 120], outline=(50, 80, 150), width=2, fill=(200, 215, 240))
    draw.polygon([(60, 50), (100, 20), (140, 50)], outline=(50, 80, 150), fill=(170, 190, 220))
    draw.polygon([(160, 70), (200, 45), (240, 70)], outline=(50, 80, 150), fill=(170, 190, 220))

    # テキスト
    draw.text((10, 5), text, fill=(30, 50, 120))
    draw.text((10, 130), "スケール 1:200", fill=(80, 80, 80))

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def create_company_logo(width=200, height=60) -> io.BytesIO:
    """簡易ロゴ画像を生成する"""
    img = Image.new("RGB", (width, height), color=(255, 255, 255))
    draw = ImageDraw.Draw(img)
    draw.rectangle([5, 5, 195, 55], outline=(30, 60, 150), width=2)
    draw.rectangle([5, 5, 70, 55], fill=(30, 60, 150))
    draw.text((15, 20), "株式会社", fill=(255, 255, 255))
    draw.text((80, 10), "サンプル建設", fill=(30, 60, 150))
    draw.text((80, 35), "SAMPLE KENSETSU Co.,Ltd.", fill=(100, 100, 100))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# シート1: 工事仕様書（典型的な方眼紙レイアウト）
# ---------------------------------------------------------------------------

def create_sheet_spec(wb: Workbook):
    ws = wb.active
    ws.title = "工事仕様書"
    set_grid(ws, max_col=50, max_row=70)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    white_bold  = Font(name="メイリオ", bold=True, color="FFFFFF", size=11)
    black_bold  = Font(name="メイリオ", bold=True, color="000000", size=10)
    black_small = Font(name="メイリオ", color="000000", size=9)
    title_font  = Font(name="メイリオ", bold=True, color="FFFFFF", size=14)

    # --- タイトルブロック (行1-3, 列1-50) ---
    merge_and_write(ws, 1, 1, 3, 50,
                    "工 事 仕 様 書（サンプル）",
                    font=title_font, fill=TITLE_FILL,
                    border=MEDIUM_BORDER,
                    alignment=center)

    # --- 基本情報ブロック ---
    labels = [
        ("工事名称",    4,  1,  5,  6,  "○○地区 新築工事",        4,  7,  5, 20),
        ("工事場所",    4,  21, 5, 26,  "東京都○○区○○町1-2-3",    4, 27,  5, 50),
        ("発注者",      6,  1,  7,  6,  "株式会社 ○○建設",          6,  7,  7, 20),
        ("施工者",      6,  21, 7, 26,  "株式会社 サンプル建設",      6, 27,  7, 50),
        ("工期",        8,  1,  9,  6,  "2025年4月1日 ～ 2026年3月31日", 8, 7, 9, 50),
        ("図面番号",   10,  1, 11,  6,  "S-001",                    10,  7, 11, 20),
        ("改訂",       10, 21, 11, 26,  "Rev.0",                    10, 27, 11, 50),
    ]
    for label, lr1, lc1, lr2, lc2, val, vr1, vc1, vr2, vc2 in labels:
        merge_and_write(ws, lr1, lc1, lr2, lc2, label,
                        font=black_bold, fill=SUBHEADER_FILL,
                        border=THIN_BORDER, alignment=center)
        merge_and_write(ws, vr1, vc1, vr2, vc2, val,
                        font=black_small, fill=None,
                        border=THIN_BORDER, alignment=left)

    # --- セクションヘッダー: 仕様一覧 ---
    merge_and_write(ws, 13, 1, 13, 50,
                    "1. 主要材料仕様",
                    font=white_bold, fill=HEADER_FILL,
                    border=THIN_BORDER, alignment=left)

    # 仕様テーブルヘッダー
    headers = [("項目", 1, 4), ("材料名", 5, 14), ("規格・品番", 15, 28),
               ("数量", 29, 33), ("単位", 34, 37), ("備考", 38, 50)]
    for h_text, c1, c2 in headers:
        merge_and_write(ws, 14, c1, 14, c2, h_text,
                        font=black_bold, fill=SUBHEADER_FILL,
                        border=THIN_BORDER, alignment=center)

    # 仕様データ
    spec_data = [
        ("1", "コンクリート",         "Fc=24N/mm²",           120, "m³",  "基礎・床"),
        ("2", "鉄筋",                 "SD345 D16",             8.5, "t",   "主筋"),
        ("3", "構造用合板",           "JAS特類 t=12mm",        340, "枚",  "床・壁下地"),
        ("4", "断熱材（グラスウール）", "HG16-105mm",           280, "m²",  "外壁充填"),
        ("5", "アルミサッシ",         "複層ガラス Low-E",       45, "箇所", "断熱仕様"),
        ("6", "屋根材",               "カラーガルバリウム鋼板 t=0.4mm", 180, "m²", ""),
        ("7", "外壁材",               "窯業系サイディング t=14mm", 310, "m²", "塗装品"),
    ]
    for i, (no, mat, spec, qty, unit, note) in enumerate(spec_data):
        row = 15 + i
        for val, c1, c2 in [(no, 1, 4), (mat, 5, 14), (spec, 15, 28),
                            (qty, 29, 33), (unit, 34, 37), (note, 38, 50)]:
            fill = LABEL_FILL if i % 2 == 0 else None
            merge_and_write(ws, row, c1, row, c2, val,
                            font=black_small, fill=fill,
                            border=THIN_BORDER,
                            alignment=center if c2 - c1 < 6 else left)

    # --- セクションヘッダー: 施工注意事項 ---
    note_row = 23
    merge_and_write(ws, note_row, 1, note_row, 50,
                    "2. 施工注意事項",
                    font=white_bold, fill=HEADER_FILL,
                    border=THIN_BORDER, alignment=left)

    notes = [
        "① 本仕様書は設計図書と合わせて使用すること。",
        "② 材料の搬入前に監督員の承認を得ること。",
        "③ 各工程完了時に写真記録を行い、施工管理台帳に添付すること。",
        "④ 寸法は原則として現場実測を優先し、疑義が生じた場合は監督員に確認すること。",
        "⑤ 廃材の処理は廃棄物処理法に従い適正に行うこと。",
    ]
    for i, note_text in enumerate(notes):
        row = note_row + 1 + i
        merge_and_write(ws, row, 1, row, 50, note_text,
                        font=black_small, fill=None,
                        border=THIN_BORDER, alignment=left)

    # --- ロゴ画像を右上に貼り付け ---
    logo_buf = create_company_logo()
    logo_img = XLImage(logo_buf)
    logo_img.width = 160
    logo_img.height = 48
    logo_img.anchor = "AH1"
    ws.add_image(logo_img)

    # --- 概要図を貼り付け ---
    merge_and_write(ws, 30, 1, 30, 50,
                    "3. 配置概要図",
                    font=white_bold, fill=HEADER_FILL,
                    border=THIN_BORDER, alignment=left)

    fig_buf = create_dummy_image(400, 200, text="配置概要図（イメージ）")
    fig_img = XLImage(fig_buf)
    fig_img.width = 400
    fig_img.height = 200
    fig_img.anchor = "A31"
    ws.add_image(fig_img)


# ---------------------------------------------------------------------------
# シート2: 工程表（ガントチャート風）
# ---------------------------------------------------------------------------

def create_sheet_schedule(wb: Workbook):
    ws = wb.create_sheet("工程表")
    set_grid(ws, max_col=55, max_row=35)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    white_bold  = Font(name="メイリオ", bold=True, color="FFFFFF", size=9)
    black_bold  = Font(name="メイリオ", bold=True, color="000000", size=9)
    black_small = Font(name="メイリオ", color="000000", size=8)
    bar_fill    = PatternFill("solid", fgColor="4472C4")
    bar2_fill   = PatternFill("solid", fgColor="ED7D31")

    # タイトル
    merge_and_write(ws, 1, 1, 2, 55,
                    "工 程 表（サンプル）　2025年4月 ～ 2026年3月",
                    font=Font(name="メイリオ", bold=True, size=12),
                    fill=TITLE_FILL,
                    border=MEDIUM_BORDER,
                    alignment=center)

    # 月ヘッダー (列7〜54 = 48列を12ヶ月で分割)
    months = ["4月", "5月", "6月", "7月", "8月", "9月",
              "10月", "11月", "12月", "1月", "2月", "3月"]
    for i, month in enumerate(months):
        c1 = 7 + i * 4
        c2 = c1 + 3
        merge_and_write(ws, 3, c1, 3, c2, month,
                        font=white_bold, fill=HEADER_FILL,
                        border=THIN_BORDER, alignment=center)

    # 作業列ヘッダー
    for text, c1, c2 in [("No.", 1, 1), ("工種", 2, 4), ("担当", 5, 6)]:
        merge_and_write(ws, 3, c1, 3, c2, text,
                        font=black_bold, fill=SUBHEADER_FILL,
                        border=THIN_BORDER, alignment=center)

    # ガントチャートデータ (開始月index, 期間月数)
    tasks = [
        ("1", "仮設工事",     "A班",  0, 2),
        ("2", "土工事・基礎", "B班",  1, 3),
        ("3", "鉄筋工事",     "B班",  3, 2),
        ("4", "型枠工事",     "C班",  3, 3),
        ("5", "コンクリート", "B班",  5, 2),
        ("6", "鉄骨工事",     "D班",  4, 4),
        ("7", "外壁工事",     "E班",  7, 3),
        ("8", "屋根工事",     "E班",  6, 2),
        ("9", "内装工事",     "F班",  8, 4),
        ("10", "設備工事",    "G班",  6, 6),
        ("11", "外構工事",    "H班", 10, 2),
        ("12", "竣工検査",    "全班", 11, 1),
    ]
    for i, (no, task, tanto, start, duration) in enumerate(tasks):
        row = 4 + i
        fill = LABEL_FILL if i % 2 == 0 else None

        for val, c1, c2 in [(no, 1, 1), (task, 2, 4), (tanto, 5, 6)]:
            merge_and_write(ws, row, c1, row, c2, val,
                            font=black_small, fill=fill,
                            border=THIN_BORDER,
                            alignment=center if c2 == c1 else left)

        # ガントバー
        bar_c1 = 7 + start * 4
        bar_c2 = bar_c1 + duration * 4 - 1
        b_fill = bar_fill if i % 3 != 2 else bar2_fill
        merge_and_write(ws, row, bar_c1, row, bar_c2, "",
                        fill=b_fill, border=THIN_BORDER, alignment=center)

        # バー外のセルに罫線
        for c in range(7, 55):
            if not (bar_c1 <= c <= bar_c2):
                ws.cell(row=row, column=c).border = THIN_BORDER
                if fill:
                    ws.cell(row=row, column=c).fill = fill


# ---------------------------------------------------------------------------
# シート3: 数量集計表 + グラフ
# ---------------------------------------------------------------------------

def create_sheet_quantity(wb: Workbook):
    ws = wb.create_sheet("数量集計表")
    set_grid(ws, max_col=40, max_row=50)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")
    white_bold  = Font(name="メイリオ", bold=True, color="FFFFFF", size=10)
    black_bold  = Font(name="メイリオ", bold=True, size=10)
    black_small = Font(name="メイリオ", size=9)

    # タイトル
    merge_and_write(ws, 1, 1, 2, 40,
                    "数 量 集 計 表（サンプル）",
                    font=Font(name="メイリオ", bold=True, size=12),
                    fill=TITLE_FILL,
                    border=MEDIUM_BORDER,
                    alignment=center)

    # ヘッダー行
    headers = [("工種", 1, 6), ("単位", 7, 10), ("設計数量", 11, 16),
               ("計画数量", 17, 22), ("実施数量", 23, 28),
               ("達成率(%)", 29, 34), ("備考", 35, 40)]
    for h, c1, c2 in headers:
        merge_and_write(ws, 3, c1, 3, c2, h,
                        font=white_bold, fill=HEADER_FILL,
                        border=THIN_BORDER, alignment=center)

    # データ
    data_rows = [
        ("コンクリート打設", "m³",  120,  122,  115),
        ("鉄筋組立",         "t",   8.5,  8.5,  8.2),
        ("型枠",             "m²",  340,  345,  338),
        ("外壁サイディング", "m²",  310,  312,  295),
        ("断熱材充填",       "m²",  280,  280,  280),
        ("屋根葺き",         "m²",  180,  180,  175),
        ("アルミサッシ取付", "箇所",  45,   45,   43),
    ]
    for i, (koujyu, unit, design, plan, actual) in enumerate(data_rows):
        row = 4 + i
        fill = LABEL_FILL if i % 2 == 0 else None
        rate = round(actual / design * 100, 1) if design else 0

        for val, c1, c2, align in [
            (koujyu, 1,  6,  left),
            (unit,   7,  10, center),
            (design, 11, 16, right),
            (plan,   17, 22, right),
            (actual, 23, 28, right),
            (rate,   29, 34, center),
            ("",     35, 40, left),
        ]:
            merge_and_write(ws, row, c1, row, c2, val,
                            font=black_small, fill=fill,
                            border=THIN_BORDER, alignment=align)

    # 合計行
    row = 4 + len(data_rows)
    merge_and_write(ws, row, 1, row, 6, "合　計 / 平均",
                    font=black_bold, fill=SUBHEADER_FILL,
                    border=THIN_BORDER, alignment=center)
    for c1, c2 in [(7, 10), (11, 16), (17, 22), (23, 28), (29, 34), (35, 40)]:
        merge_and_write(ws, row, c1, row, c2, "",
                        font=black_bold, fill=SUBHEADER_FILL,
                        border=THIN_BORDER, alignment=center)

    # --- グラフ（棒グラフ: 設計 vs 実施数量）---
    # グラフ用データ（数値のみ、シート上の非表示列に配置）
    chart_data_col = 42
    ws.cell(row=3, column=chart_data_col, value="工種")
    ws.cell(row=3, column=chart_data_col + 1, value="設計数量")
    ws.cell(row=3, column=chart_data_col + 2, value="実施数量")

    chart_labels = []
    for i, (koujyu, unit, design, plan, actual) in enumerate(data_rows):
        r = 4 + i
        ws.cell(row=r, column=chart_data_col, value=koujyu)
        ws.cell(row=r, column=chart_data_col + 1, value=design)
        ws.cell(row=r, column=chart_data_col + 2, value=actual)
        chart_labels.append(koujyu)

    chart = BarChart()
    chart.type = "col"
    chart.title = "設計数量 vs 実施数量"
    chart.y_axis.title = "数量"
    chart.x_axis.title = "工種"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref = Reference(ws, min_col=chart_data_col + 1, max_col=chart_data_col + 2,
                         min_row=3, max_row=3 + len(data_rows))
    cats = Reference(ws, min_col=chart_data_col, min_row=4,
                     max_row=3 + len(data_rows))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A13")

    # --- 追加画像（断面図イメージ） ---
    sec_buf = create_dummy_image(350, 150, text="断面図（サンプル）")
    sec_img = XLImage(sec_buf)
    sec_img.width = 350
    sec_img.height = 150
    sec_img.anchor = "A35"
    ws.add_image(sec_img)


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------

def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    wb = Workbook()
    create_sheet_spec(wb)
    create_sheet_schedule(wb)
    create_sheet_quantity(wb)

    wb.save(OUTPUT_FILE)
    print(f"テストデータを生成しました: {OUTPUT_FILE}")
    print(f"  シート数: {len(wb.sheetnames)}")
    for sn in wb.sheetnames:
        print(f"    - {sn}")


if __name__ == "__main__":
    main()
